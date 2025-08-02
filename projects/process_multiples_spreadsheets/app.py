import openpyxl as excel
from openpyxl.chart import BarChart, Reference

def process_workbook(workbook):
    workbook = excel.load_workbook(workbook)
    sheet_name = input("Ingrese el nombre de la hoja que desea leer (Tener en cuenta las mayusculas y minusculas): ")
    sheet = None
    try:
        sheet = workbook[sheet_name]  # Verifica si la hoja existe.
    except KeyError:
        print(f"La hoja '{sheet_name}' no existe en el libro de trabajo.")
        exit()

    for row in range(2, sheet.max_row + 1):
        price_cell = sheet.cell(row, 3)
        corrected_price = price_cell.value * 0.90
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price  # Se asigna el valor corregido a la celda correspondiente.

    new_cell = input("Indique el nombre de la columna nueva donde quedaran los precios corregidos: ")
    sheet.cell(1, 4).value = new_cell  # Se asigna el nombre de la nueva columna en la primera fila.

    # Se crea un gráfico de barras para visualizar los precios corregidos.
    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart = BarChart()
    chart.add_data(values)
    chart.title = new_cell
    sheet.add_chart(chart, "E2") # Se agrega el gráfico a la hoja en la celda A5.

    workbook.save(workbook)  # Se guarda el libro de trabajo con los cambios realizados.
    workbook.close()  # Se cierra el libro de trabajo para liberar recursos.
    print("El archivo 'transactions_corrected.xlsx' ha sido creado con los precios corregidos.")

if __name__ == "__main__":
    try:
        workbook_file = input("Ingrese el nombre del archivo de Excel (con extensión .xlsx): ")
        process_workbook(workbook_file)
    except FileNotFoundError:
        print("El archivo 'transactions.xlsx' no se encuentra en el directorio actual.")
    except Exception as e:
        print(f"Ocurrió un error: {e}")